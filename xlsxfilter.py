import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Ask the user for the Excel file name
file_name = input("Please enter the name of the Excel file: ")

# Load the Excel file
df = pd.read_excel(file_name)

# Ask the user for the license type
license_type = input("Are you filtering for 'Standard' or 'Basic' licenses? ")
license_text = "Microsoft 365 Business " + license_type

# Define a function to check if the license text is in a row
def check_row(row):
    return license_text in row.values

# Apply the function to each row
mask = df.apply(check_row, axis=1)

# Keep only the rows that contain the license text
df = df[mask]

# Create the output file name
base_name = os.path.basename(file_name)
name, ext = os.path.splitext(base_name)
output_file_name = license_type.lower() + "_" + name + ext

# Write the result back to the new Excel file
df.to_excel(output_file_name, index=False)

# Load the workbook
wb = load_workbook(output_file_name)

# Select the active worksheet
ws = wb.active

# Iterate over the columns
for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

# Save the workbook
wb.save(output_file_name)
